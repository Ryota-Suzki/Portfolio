import glob
import os
import pandas as pd
import re
from datetime import datetime as dt, date, timezone, timedelta
import pathlib
from openpyxl import load_workbook
import datetime

#変数
path_join = ""
col_names = [ 'Data{0:02d}'.format(i) for i in range(51) ] #Colums番号

path = pathlib.Path().absolute() #Excel-Path
path /='../../../●●●●●●/'
# print(path.resolve()) #デバッグ
path = str(path)

FILE_PATH = glob.glob("●●●●●●*/●●●●●●.*") #gclog-Path

file_creatiing_years = path_join.join((re.findall("●●●●●●(.*)",FILE_PATH[0].split(os.sep)[0])))[0:4]
creation_date = datetime.datetime.strptime(path_join.join((re.findall("●●●●●●(.*)",FILE_PATH[0].split(os.sep)[0]))),'%Y-%m-%d')
print("■ファイル年月日「" + str(creation_date) + "」")


if file_creatiing_years not in (path_join.join(glob.glob(path + "Excelファイル名" + file_creatiing_years + ".xlsx"))):
    print("ERROR"+"//ファイルが存在するかもしくはファイル名の年月日が正しいか確認してください。//")
    exit()
else:
    PATH =  path_join.join(glob.glob(path + "Excelファイル名" + file_creatiing_years + ".xlsx"))

# 関数

def gclog_shaping(File): # データ整形

    global gclog_value
    global gclog_datetime

    data = pd.read_csv(File,encoding='cp932', sep='\s',names=col_names,engine="python")
    data = data[data["Data07"] == "[ParNew:"]
    gclog_time = (data["Data00"]).tail(1)

    gclog_list = gclog_time.astype(str).tolist()
    gclog_list = path_join.join(gclog_list)

    gclog_day = (gclog_list[:10])
    gclog_time = (gclog_list[11:19])
    gclog_dt = gclog_day + " " + gclog_time

    gclog_datetime = dt.strptime(gclog_dt,'%Y-%m-%d %H:%M:%S')

    gclog_num = data["Data11"].tail(1)
    gclog_list = gclog_num.astype(str).tolist()

    gclog_value = path_join.join(gclog_list)
    gclog_value = gclog_value.partition("(")[0]
    gclog_value = gclog_value.replace("K->"," ").replace("K","")

# 書き込みシート選択処理用関数
def sheet_celection(cell_num):
    global CellAddrs,CellValue

    CellAddrs = []
    CellValue = []

    for cell in sheet[cell_num]:
        CellAddrs.append(cell.coordinate)
        CellValue.append(str(cell.value))

# 書き込みシート＆セル指定
def javaheap_excelfile_write(path):

    global javaheap_num
    global wb
    global sheet

    wb = load_workbook(path,data_only=True)

    sheets = wb.sheetnames
    sheet = wb[sheets[0]]

    sheet_celection("B")# 書き込みシート選択処理用関数

    if str(creation_date) not in CellValue:
        sheet = wb[sheets[1]]
        sheet_celection("B")# 書き込みシート選択処理用関数

    if str(creation_date) not in CellValue:
        sheet = wb[sheets[2]]
        sheet_celection("B")# 書き込みシート選択処理用関数

    if str(creation_date) not in CellValue:
        sheet = wb[sheets[3]]
        sheet_celection("B")# 書き込みシート選択処理用関数

    if str(creation_date) not in CellValue:
        print("ERROR" + "//セル内の年月日が正しく入力されているか確認してください。//")
        exit() #

    df = pd.DataFrame({"Num":CellAddrs,"Value":CellValue})
    df = (df.query('Value.str.contains("{}")'.format(creation_date)))

    javaheap_num = df["Num"].astype(str).tolist()
    javaheap_num = path_join.join(javaheap_num)

#メイン処理
print(str(len(FILE_PATH)) + "個のファイルが存在します。")
#ファイル数が日によって異なる為、個数によっての処理実装


if len(FILE_PATH) == 1: #ファイル数が"1"の場合
    File = path_join.join(FILE_PATH[0])
    print("■ Read files")

    gclog_shaping(File) # データ整形
    javaheap_excelfile_write(PATH) # 書き込みシート＆セル指定
    javaheap_cell_before = javaheap_num.replace("B","C")

    gclog_datetime = gclog_datetime
    gclog_Before = gclog_value.partition(" ")[0]
    gclog_After = gclog_value.partition(" ")[2]


    javaheap_excelfile_write(PATH)  # 書き込みシート＆セル指定
    print("選択シート" + str(sheet))
    javaheap_cell_after = javaheap_num.replace("B","D")

    sheet[javaheap_cell_before] = int(gclog_Before)
    sheet[javaheap_cell_after] = int(gclog_After)

    print("GC前：" + str(gclog_Before))
    print("GC後：" + str(gclog_After))

    print("\n" + "ファイルを正常に処理しました。")

    wb.save(filename= PATH)

elif len(FILE_PATH) == 2: #ファイル数が"2"の場合(データ整形部分は関数に置き換え予定)
    File01 = path_join.join(FILE_PATH[0])
    File02 = path_join.join(FILE_PATH[1])
    print("■ Read Two Files")  #デバッグ

    gclog_shaping(File01) # データ整形
    gclog01_datetime = gclog_datetime
    gclog01_Before = gclog_value.partition(" ")[0]
    gclog01_After = gclog_value.partition(" ")[2]


    gclog_shaping(File02) # データ整形
    gclog02_datetime = gclog_datetime
    gclog02_Before = gclog_value.partition(" ")[0]
    gclog02_After = gclog_value.partition(" ")[2]

    javaheap_excelfile_write(PATH)  # 書き込みシート＆セル指定
    print("選択シート" + str(sheet))
    javaheap_cell_before = javaheap_num.replace("B","C")


    javaheap_excelfile_write(PATH)  # 書き込みシート＆セル指定
    javaheap_cell_after = javaheap_num.replace("B","D")


    if gclog01_datetime > gclog02_datetime:
        print(str(FILE_PATH[0]) + ":" + "抽出")
        print(sheet)
        print(sheet[javaheap_cell_before])
        sheet[javaheap_cell_before] = int(gclog01_Before)
        sheet[javaheap_cell_after] = int(gclog01_After)
        print("GC前：" + str(gclog01_Before))
        print("GC後：" + str(gclog01_After))
    else:
        print(str(FILE_PATH[1]) + ":" + "抽出")
        sheet[javaheap_cell_before] = int(gclog02_Before)
        sheet[javaheap_cell_after] = int(gclog02_After)
        print("GC前：" + str(gclog02_Before))
        print("GC後：" + str(gclog02_After))

    print("\n" + "複数ファイルを正常に処理しました。")

    wb.save(filename= PATH)

elif len(FILE_PATH) == 3: #ファイル数が"3"の場合(データ整形部分は関数に置き換え予定)
    File01 = path_join.join(FILE_PATH[0])
    File02 = path_join.join(FILE_PATH[1])
    File03 = path_join.join(FILE_PATH[2])
    print("■ Read Three Files") #デバッグ

    gclog_shaping(File01) # データ整形
    gclog01_datetime = gclog_datetime
    gclog01_Before = gclog_value.partition(" ")[0]
    gclog01_After = gclog_value.partition(" ")[2]

    gclog_shaping(File02) # データ整形
    gclog02_datetime = gclog_datetime
    gclog02_Before = gclog_value.partition(" ")[0]
    gclog02_After = gclog_value.partition(" ")[2]

    gclog_shaping(File03) # データ整形
    gclog03_datetime = gclog_datetime
    gclog03_Before = gclog_value.partition(" ")[0]
    gclog03_After = gclog_value.partition(" ")[2] ##今後関数に置き換え

    javaheap_excelfile_write(PATH)  # 書き込みシート＆セル指定
    print("選択シート" + str(sheet))
    javaheap_cell_before = javaheap_num.replace("B","C")

    javaheap_excelfile_write(PATH)  # 書き込みシート＆セル指定
    javaheap_cell_after = javaheap_num.replace("B","D")

    print("\n" + "複数ファイルを正常に処理しました。")


    if gclog01_datetime > gclog03_datetime and gclog01_datetime > gclog02_datetime:
        print(str(FILE_PATH[0]) + ":" + "抽出")
        sheet[javaheap_cell_before] = int(gclog01_Before)
        sheet[javaheap_cell_after] = int(gclog01_After)
        print("GC前：" + str(gclog01_Before))
        print("GC前：" + str(gclog01_After))

    elif gclog02_datetime > gclog03_datetime and gclog02_datetime > gclog01_datetime:
        print(str(FILE_PATH[1]) + ":" + "抽出")
        sheet[javaheap_cell_before] = int(gclog02_Before)
        sheet[javaheap_cell_after] = int(gclog02_After)
        print("GC前：" + str(gclog02_Before))
        print("GC前：" + str(gclog02_After))

    elif gclog03_datetime > gclog01_datetime and gclog03_datetime > gclog02_datetime:
        print(str(FILE_PATH[2]) + ":" + "抽出")
        sheet[javaheap_cell_before] = int(gclog03_Before)
        sheet[javaheap_cell_after] = int(gclog03_After)
        print("GC前：" + str(gclog03_Before))
        print("GC前：" + str(gclog03_After))

    else:
        print("ERROR")

    wb.save(filename= PATH)
