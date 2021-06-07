# -*- coding: utf-8 -*-
import glob
import pandas as pd
import openpyxl as op
import re
import os
import datetime
import pathlib

from openpyxl import Workbook, load_workbook
from pathlib import Path

path_join = ""

def sheet_celection(): #シート判定用関数
    global cell_value,cell_addrs
    cell_value=[]
    cell_addrs=[]

    for i in range(1,1000): #列番号1~1000のcell_valueとセル番号を抽出し、リストに追加
        cell_value.append(str(detection_sheet.cell(row = 4,column = i).value))
        cell_addrs.append(str(detection_sheet.cell(row = 4,column = i).coordinate))

def excel_set(name,file_year,file_creation_date): #Excelファイル指定用関数
    global dbcp_workbook, dbcp_sheet, detection_sheet, creation_date ,detection_excel_path, dbcp_excel_path

    path = pathlib.Path().absolute()
    path /= '../../../{0}/'.format(name) # ディレクトリ移動

    if name == "DBCP":

        # print("対象ファイルパス:" + path) #デバッグ
        dbcp_excel_path = path_join.join(glob.glob(str(path) + "Excelファイル名" + file_year + ".xlsx"))
        print("\n" + "■対象Excelファイル\n" + "ブック:" + dbcp_excel_path)

        workbook_for_writing = load_workbook(dbcp_excel_path)
        workbook_for_writing.save(dbcp_excel_path)

        dbcp_workbook = load_workbook(dbcp_excel_path) #書き込みセル指定用
        dbcp_sheet = dbcp_workbook[path_join.join(dbcp_workbook.sheetnames)]

        print("シート名：" + str(dbcp_sheet) + "\n")

    elif name == "検知":

        detection_excel_path = path_join.join(glob.glob(str(path) + "Excelファイル名" + file_year + ".xlsx"))
        print("ブック:" + detection_excel_path)

        workbook_for_extracting = load_workbook(detection_excel_path,data_only=True) #書き込みセル指定用
        workbook_for_extracting.save(detection_excel_path)
        sheets_for_extracting = workbook_for_extracting.sheetnames #書き込みセル指定用

        #Sheet設定
        detection_sheet = workbook_for_extracting[sheets_for_extracting[0]]
        sheet_celection()

        if str(file_creation_date) not in cell_value:
            detection_sheet = workbook_for_extracting[sheets_for_extracting[1]]
            sheet_celection()

        if str(file_creation_date) not in cell_value:
            detection_sheet = workbook_for_extracting[sheets_for_extracting[2]]
            sheet_celection()

        if str(file_creation_date) not in cell_value:
            detection_sheet = workbook_for_extracting[sheets_for_extracting[3]]
            sheet_celection()

        elif str(file_creation_date) not in cell_value:
            print("ERROR" + "//セル内の年月日が正しく入力されているか確認してください。//")
            exit()

        print("シート名：" + str(detection_sheet) + "\n")
        creation_date = file_creation_date

def detection_excel_write(unit_name, in_time_dbcp, out_time_dbcp): #Excelファイル書き込み用
    global confirm_cell
    cell_value=[]
    cell_addrs=[]

    workbook_for_writing =  load_workbook(detection_excel_path)
    sheet_for_writing = workbook_for_writing[detection_sheet.title]

    for i in range(1,1000): #列番号1~1000のcell_valueとセル番号を抽出し、リストに追加
        cell_value.append(str(detection_sheet.cell(row = 4,column = i).value))
        cell_addrs.append(str(detection_sheet.cell(row = 4,column = i).coordinate))

    df = pd.DataFrame({"Num":cell_addrs,"Value":cell_value}) #上記リストをDataFrameに格納
    df = (df.query('Value.str.contains("{}")'.format(creation_date)))
    confirm_cell = path_join.join(df["Num"].astype(str).tolist()).replace("4","8") #リスト変換

    if unit_name == "●●●●●●":
        sheet_for_writing[confirm_cell] = in_time_dbcp
        sheet_for_writing[confirm_cell.replace("8","9")] = out_time_dbcp

    elif unit_name == "●●●●●●":
        sheet_for_writing[confirm_cell.replace("8","10")] = in_time_dbcp
        sheet_for_writing[confirm_cell.replace("8","11")] = out_time_dbcp

    workbook_for_writing.save(filename = detection_excel_path)

def dbcp_excel_write(log_data, start_date, unit_name, alart_count): #Excelファイル書き込み用
    global confirm_cell
    cell_value=[]
    cell_addrs=[]

    if unit_name == "●●●●●●":
        workbook_for_writing =  load_workbook(dbcp_excel_path)
        sheet_for_writing = workbook_for_writing[dbcp_sheet.title]

        for cell in dbcp_sheet["B"]:
            cell_value.append(str(cell.value))
            cell_addrs.append(str(cell.coordinate))

        df = pd.DataFrame({"Num":cell_addrs,"Value":cell_value}) #上記リストをDataFrameに格納
        df = (df.query('Value.str.contains("{}")'.format(creation_date)))

        confirm_cell = path_join.join(df["Num"].astype(str).tolist()).replace("B","D") #リスト変換

        if (log_data.iat[0,0]) > start_date:
            sheet_for_writing[confirm_cell] = min(log_data["DBCP"], key=int)
            sheet_for_writing[confirm_cell.replace("D","E")] = max(log_data["DBCP"], key=int)

        elif (log_data.iat[0,0]) < start_date:
            sheet_for_writing[confirm_cell.replace("D","F")] = min(log_data["DBCP"], key=int)
            sheet_for_writing[confirm_cell.replace("D","G")] = max(log_data["DBCP"], key=int)

        workbook_for_writing.save(filename = dbcp_excel_path)

    else:
        pass

    if len(alart_count) > 0 and (alart_count.iat[0,1]) == "●●●●●●": #alart_count[0,1]に該当文字列が存在するか.

        alart_time = alart_count["Date"]
        alart_dbcp = alart_count["DBCP"].values.tolist()

        alart_list = []
        for alart_time, alart_dbcp in zip(alart_time, alart_dbcp):
            new_alart_date, new_alart_time = str(alart_time).split() #アラートが発生してる日時を分割
            alart_list.append(str(new_alart_time) + " " + str(alart_dbcp) + "\n") #閾値を超えたDBCPと発生日時を結合し、リストに追加
            alart_data = path_join.join(alart_list)
        print(alart_data)

        for count in range(len(alart_list)):
            sheet_for_writing[confirm_cell.replace("D","H")] = str(alart_data.rstrip("\r\n"))

        workbook_for_writing.save(filename = dbcp_excel_path)

    elif len(alart_count) > 0 and (alart_count.iat[0,1]) == "●●●●●●":
        print("DBCP数が30を越えています。ログを確認してください。")

    else:
        pass
