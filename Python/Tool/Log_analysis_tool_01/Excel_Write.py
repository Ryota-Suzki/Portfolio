# -*- coding: utf-8 -*-


import glob
import pandas as pd
import openpyxl as op
from openpyxl import Workbook, load_workbook
import pathlib
import re
import os

path_join = ","

def excel_dbcp_writing(file_creating_years,file_creation_date,one_unit_deadlock_count): #作成年、作成日、デッドロック数
    path = pathlib.Path().absolute()   # Main.pyのあるディレクトリ
    path /= '../../../●●●●●●/'     # ディレクトリ移動　Main.pyの階層から3つ上の対象フォルダを指定

    dir_location = str(path)
    print(dir_location)

    excel_path = path_join.join(glob.glob(dir_location + "Excelファイル名" + file_creating_years + ".xlsx")) #PATH変数

    workbook_for_extracting = load_workbook(excel_path) #書き込みセル判定用
    sheets_for_extracting = workbook_for_extracting[path_join.join(workbook_for_extracting.sheetnames)] #書き込みセル判定用

    cell_addrs = []
    cell_value =[]

    for cell in sheets_for_extracting["B"]:
        cell_addrs.append(cell.coordinate) #セル番号をリストに格納
        cell_value.append(str(cell.value)) #セルの中身をリストに格納

    df = pd.DataFrame({"Num":cell_addrs,"Value":cell_value}) #上記リストをDataFrameに格納
    df = (df.query('Value.str.contains("{}")'.format(file_creation_date))) #DataFrame内を作成日で検索

    confirm_cell = path_join.join(df["Num"].astype(str).tolist()).replace("B","C") #一致したセル番号を"B"から"C"に変更し、リスト変換

    workbook_for_writing = load_workbook(excel_path) #書き込み用
    sheet_for_writing = workbook_for_writing[path_join.join(workbook_for_writing.sheetnames)] #書き込み用

    sheet_for_writing[confirm_cell] = one_unit_deadlock_count #一致したセル番号にデッドロック数を入力

    workbook_for_writing.save(filename = excel_path) #Excel保存


def sheet_celection(): #シート選択関数
    global cell_value,cell_addrs
    cell_value=[]
    cell_addrs=[]

    for i in range(1,1000): #列番号1~1000のセルの中身とセル番号を抽出し、リストに追加
        cell_value.append(str(sheet.cell(row = 4,column = i).value))
        cell_addrs.append(str(sheet.cell(row = 4,column = i).coordinate))

def excel_detec_alarm_writing(file_creating_years,file_creation_date,one_unit_deadlock_count,two_unit_deadlock_count):
    global sheet,excel_path

    path = pathlib.Path().absolute()
    path /= '../../../●●●●●●/'

    dir_location = str(path)

    excel_path = path_join.join(glob.glob(dir_location + "Excelファイル名" + file_creating_years + ".xlsx")) #PATH変数

    workbook_for_extracting = load_workbook(excel_path,data_only=True) #書き込みセル判定用
    workbook_for_extracting.save(excel_path)
    sheets_for_extracting = workbook_for_extracting.sheetnames #書き込みセル判定用

    #Sheet設定(複数シートがある場合)
    sheet = workbook_for_extracting[sheets_for_extracting[0]]
    sheet_celection()

    if str(file_creation_date) not in cell_value: #ファイル作成日がシート0のセルの中身(1~1000)と一致するか？しない場合、次のシートへ!
        sheet = workbook_for_extracting[sheets_for_extracting[1]]
        sheet_celection()

    if str(file_creation_date) not in cell_value:
        sheet = workbook_for_extracting[sheets_for_extracting[2]]
        sheet_celection()

    if str(file_creation_date) not in cell_value:
        sheet = workbook_for_extracting[sheets_for_extracting[3]]
        sheet_celection()

    elif str(file_creation_date) not in cell_value:
        print("ERROR" + "//セル内の年月日が正しく入力されているか確認してください。//")
        exit()

    print(sheet)

    writer("one_unit",file_creation_date,one_unit_deadlock_count,two_unit_deadlock_count)
    writer("two_unit",file_creation_date,one_unit_deadlock_count,two_unit_deadlock_count)

# 書き込み関数
def writer(name,file_creation_date,one_unit_deadlock_count,two_unit_deadlock_count):
    global workbook_for_writing
    cell_value=[]
    cell_addrs=[]

    workbook_for_writing =  load_workbook(excel_path)
    sheet_for_writing = workbook_for_writing[sheet.title]

    for i in range(1,1000): #列番号1~1000のcell_valueとセル番号を抽出し、リストに追加
        cell_value.append(str(sheet.cell(row = 4,column = i).value))
        cell_addrs.append(str(sheet.cell(row = 4,column = i).coordinate))

    df = pd.DataFrame({"Num":cell_addrs,"Value":cell_value}) #上記リストをDataFrameに格納
    df = (df.query('Value.str.contains("{}")'.format(file_creation_date)))

    if name == "one_unit":
        confirm_cell = path_join.join(df["Num"].astype(str).tolist()).replace("4","5") #リスト変換
        sheet_for_writing[confirm_cell] = one_unit_deadlock_count

    elif name == "two_unit":
        confirm_cell = path_join.join(df["Num"].astype(str).tolist()).replace("4","6") #リスト変換
        sheet_for_writing[confirm_cell] = two_unit_deadlock_count

    else:
        print("ERROR")

    workbook_for_writing.save(filename = excel_path)
