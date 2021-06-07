import re
import pandas as pd
import glob
import sys
from openpyxl import load_workbook
import pathlib
import datetime
import os
from datetime import datetime as dt, date, timezone, timedelta
import openpyxl as op

path_join = "" #Listから文字列に変換時に使用

path = pathlib.Path().absolute()   # DataCount.pyのあるディレクトリ
path /= '../../../●●●●●●/'     # ディレクトリ移動

PATH = str(path)

# "●●●●●●*"の名前を検索
for File in glob.glob("●●●●●●*/*●●●●●●*"):
    file_creatiing_years = path_join.join((re.findall("●●●●●●(.*)",File.split(os.sep)[1])))[0:4]
    creation_date = str(datetime.datetime.strptime(path_join.join(re.findall("●●●●●●(.*)",File.split(os.sep)[1])),'%Y%m%d'))
    print(file_creatiing_years,creation_date)
    # ●●●●●●*を開く
    with open(File, 'r') as data:
        data_file = data.read()

# []の任意の1文字と＝にマッチしたものを抽出
formating_data = (re.findall("●●●●●●.*[abcdef]=(.*)\t", data_file)) #*//データ件数//*

excel_path = path_join.join(glob.glob(PATH + "Excelファイル名" + file_creatiing_years + ".xlsx"))

workbook_for_extracting = load_workbook(excel_path,data_only=True) #日付判定用(*//data_only=Trueによって数式の結果を取得//*)
sheet_for_extracting = workbook_for_extracting[file_creatiing_years + "●●●●●●"] #日付判定用

CellValue = []
CellAddrs = []

for i in range(1,1000): #列番号1~1000のCellValueとセル番号を抽出し、リストに追加
    CellValue.append(str(sheet_for_extracting.cell(row = 4,column = i).value))
    CellAddrs.append(str(sheet_for_extracting.cell(row = 4,column = i).coordinate))

cell_infomation = pd.DataFrame({"Num":CellAddrs,"Value":CellValue}) #上記リストをDataFrameに格納

matched_cell = (cell_infomation .query('Value.str.contains("{}")'.format(creation_date))) #ファイルの日付に一致するか判定

cell_num = path_join.join(matched_cell["Num"].astype(str).tolist()).replace("4","5")

factor_converting = op.utils.column_index_from_string(cell_num.replace("5","")) #Openpyxl機能(セル列名を番号に変換　*セル番号の5を空白に置換*)

excel_for_writing = load_workbook(excel_path) #ExcelFile設定
sheet_for_writing = excel_for_writing[file_creatiing_years + "●●●●●●"] #sheet_for_extracting設定

row_num = int(re.sub("\\D","",cell_num)) #セル番号の数字部分だけ抽出(\D,""：10進数でない任意の文字を空白に置換) 例) S5 → 5

for statl in formating_data: #データ件数のlistをformating_dataの中身だけ順番に書き込み
    sheet_for_writing.cell(row=row_num,column=factor_converting).value = int(statl)
    row_num += 1

excel_for_writing.save(filename=excel_path)

print("正常に処理しました")
